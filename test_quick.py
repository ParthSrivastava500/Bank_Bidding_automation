"""
QUICK TEST — 3 keywords, Mumbai only, 2 pages max, no delays
Saves to Desktop/Test_Screenshots/
"""
import os, time, datetime, glob, tempfile, shutil
from pathlib import Path

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from PIL import Image, ImageDraw, ImageFont
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io

# ── TEST CONFIG ──────────────────────────────────────────────
TEST_KEYWORDS = [
    "hdfc credit card apply offer",
    "hdfc credit card apply link",
    "hdfc credit card",
]
TEST_CITY     = "Mumbai"
MAX_PAGES     = 2

CHROME_EXE        = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
CHROMEDRIVER_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "chromedriver.exe")
AWR_EXT_PATH      = r"C:\Users\sriva\AppData\Local\Google\Chrome\User Data\Profile 11\Extensions\kjffabibegojdmknibcophlonmadkpin\1.2_0"

DESKTOP_PATH = Path(os.path.expanduser("~/Desktop"))
SS_FOLDER    = DESKTOP_PATH / "Test_Screenshots"
OUTPUT_EXCEL = DESKTOP_PATH / "test_results.xlsx"
SS_COUNTER   = [1]  # mutable counter shared across calls

STOPWORDS = {"apply","online","card","get","now","for","a","the","and","in",
             "of","to","how","best","top","free","new","with","my","is","on"}

# ── EXCEL ────────────────────────────────────────────────────
def init_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Results"
    headers = ["Date","Time","keyword","location","ad_heading",
               "ad_description","source_url","destination_url","screenshot"]
    hfill  = PatternFill("solid", start_color="1F4E79")
    hfont  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill, c.font, c.alignment = hfill, hfont, halign
    for i, w in enumerate([12,10,30,12,45,55,60,60,30], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    wb.save(OUTPUT_EXCEL)
    return wb, ws

def append_row(ws, wb, kw, city, heading, desc, src, dst, path, dt):
    row  = ws.max_row + 1
    fill = PatternFill("solid", start_color="DCE6F1" if row%2==0 else "FFFFFF")
    for col, val in enumerate([
        dt.strftime("%d/%m/%Y"), dt.strftime("%H:%M:%S"),
        kw, city, heading, desc, src, dst, str(path)
    ], 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(OUTPUT_EXCEL)

# ── SCREENSHOT ───────────────────────────────────────────────
def add_watermark(img, dt):
    draw = ImageDraw.Draw(img)
    text = dt.strftime("%d/%m/%Y  %H:%M:%S")
    try:
        font = ImageFont.truetype("C:/Windows/Fonts/arial.ttf", 28)
    except Exception:
        font = ImageFont.load_default()
    bbox = draw.textbbox((0,0), text, font=font)
    tw, th = bbox[2]-bbox[0], bbox[3]-bbox[1]
    x, y = img.width - tw - 15, img.height - th - 15
    draw.text((x+2,y+2), text, font=font, fill=(0,0,0))
    draw.text((x,y), text, font=font, fill=(255,255,0))
    return img

def save_merged(b1, b2, dt):
    i1 = Image.open(io.BytesIO(b1)).convert("RGB")
    i2 = Image.open(io.BytesIO(b2)).convert("RGB")
    h  = max(i1.height, i2.height)
    if i1.height != h: i1 = i1.resize((i1.width, h), Image.LANCZOS)
    if i2.height != h: i2 = i2.resize((i2.width, h), Image.LANCZOS)
    merged = Image.new("RGB", (i1.width+i2.width, h), (255,255,255))
    merged.paste(i1,(0,0)); merged.paste(i2,(i1.width,0))
    merged = add_watermark(merged, dt)
    out_path = SS_FOLDER / f"{SS_COUNTER[0]}.jpg"
    SS_COUNTER[0] += 1
    merged.save(out_path, "JPEG", quality=90)
    print(f"    🖼️  Saved → {out_path.name}")
    return out_path

# ── BROWSER ──────────────────────────────────────────────────
def create_driver():
    options = Options()
    options.binary_location = CHROME_EXE
    tmp_dir = tempfile.mkdtemp(prefix="chrome_test_")
    options.add_argument(f"--user-data-dir={tmp_dir}")
    if os.path.exists(AWR_EXT_PATH):
        options.add_argument(f"--load-extension={AWR_EXT_PATH}")
    options.add_argument("--start-maximized")
    options.add_argument("--no-first-run")
    options.add_argument("--no-default-browser-check")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    service = Service(CHROMEDRIVER_PATH)
    driver  = webdriver.Chrome(service=service, options=options)
    driver.set_page_load_timeout(30)
    driver._tmp_dir = tmp_dir
    return driver

# ── ADS ──────────────────────────────────────────────────────
def is_relevant(ad, keyword):
    brand_words = [w for w in keyword.lower().split()
                   if w not in STOPWORDS and len(w) > 2]
    combined = (ad["heading"] + " " + ad["source_url"]).lower()
    return any(w in combined for w in brand_words)

def get_desc(container, heading):
    for sel in [".yDYNvb","div.yDYNvb",".MUxGbd","[data-sncf='1']",".lEBKkf"]:
        try:
            el = container.find_element(By.CSS_SELECTOR, sel)
            t  = el.text.strip()
            if t and len(t) > 20 and t != heading and "http" not in t:
                return t
        except Exception:
            pass
    try:
        for el in container.find_elements(By.XPATH, ".//div | .//span"):
            try:
                t = el.text.strip()
                if len(t) > 25 and t != heading and "http" not in t:
                    if not el.find_elements(By.XPATH, "./div"):
                        return t[:300]
            except Exception:
                pass
    except Exception:
        pass
    return ""

def find_ads(driver, keyword):
    ads = []
    for lbl in driver.find_elements(By.XPATH,
            "//span[text()='Sponsored'] | //div[text()='Sponsored']"):
        try:
            container = None
            for xp in ["ancestor::div[@data-hveid][1]","ancestor::div[.//h3][1]"]:
                try: container = lbl.find_element(By.XPATH, xp); break
                except Exception: pass
            if not container: continue
            for h3 in (container.find_elements(By.CSS_SELECTOR,"h3") or
                       container.find_elements(By.CSS_SELECTOR,"[role='heading']")):
                try:
                    heading = h3.text.strip()
                    if not heading or heading == "Sponsored": continue
                    try: link_el = h3.find_element(By.XPATH,"ancestor::a[1]")
                    except Exception: link_el = container.find_element(By.CSS_SELECTOR,"a[href]")
                    href = link_el.get_attribute("href") or ""
                    desc = get_desc(container, heading)
                    ad = {"heading":heading,"description":desc,
                          "source_url":href,"link_el":link_el}
                    if is_relevant(ad, keyword) and not any(a["heading"]==heading for a in ads):
                        ads.append(ad)
                except Exception: pass
        except Exception: pass
    return ads

# ── RUN ──────────────────────────────────────────────────────
def run_test():
    SS_FOLDER.mkdir(parents=True, exist_ok=True)
    wb, ws = init_excel()
    driver = None
    tmp_dir = None
    total  = 0

    print("="*55)
    print("  QUICK TEST — 3 keywords, Mumbai, 2 pages, no delays")
    print(f"  Screenshots → {SS_FOLDER}")
    print(f"  Excel       → {OUTPUT_EXCEL}")
    print("="*55)

    try:
        driver  = create_driver()
        tmp_dir = getattr(driver, "_tmp_dir", None)
        time.sleep(2)
        print("  ✅ Browser ready\n")

        for kw in TEST_KEYWORDS:
            print(f"\n🔍 Keyword: '{kw}'")
            try:
                driver.get("https://www.google.co.in/")
                time.sleep(2)
                try: driver.find_element(By.TAG_NAME,"body").send_keys(Keys.ESCAPE)
                except Exception: pass
                for sel in ["button#L2AGLb","button[aria-label*='Accept']"]:
                    try: driver.find_element(By.CSS_SELECTOR,sel).click(); time.sleep(1); break
                    except Exception: pass

                # Check CAPTCHA
                if "sorry" in driver.current_url:
                    print("  ⚠️  CAPTCHA! Please solve it manually in the browser, then press Enter here...")
                    input()

                sb = WebDriverWait(driver,10).until(
                    EC.presence_of_element_located((By.NAME,"q")))
                sb.clear(); sb.send_keys(kw); sb.send_keys(Keys.RETURN)
                time.sleep(3)

                for page_num in range(1, MAX_PAGES+1):
                    print(f"  📄 Page {page_num}...")
                    ads = find_ads(driver, kw)
                    if ads:
                        print(f"  🎯 {len(ads)} ad(s) found")
                        for idx, ad in enumerate(ads):
                            try:
                                driver.execute_script(
                                    "arguments[0].scrollIntoView({block:'center'});",
                                    ad["link_el"])
                                time.sleep(1)
                                serp_bytes = driver.get_screenshot_as_png()
                                driver.execute_script(
                                    "arguments[0].setAttribute('target','_blank');",
                                    ad["link_el"])
                                ad["link_el"].click(); time.sleep(4)
                                if len(driver.window_handles) > 1:
                                    driver.switch_to.window(driver.window_handles[-1])
                                    time.sleep(2)
                                dest_url   = driver.current_url
                                land_bytes = driver.get_screenshot_as_png()
                                driver.close()
                                driver.switch_to.window(driver.window_handles[0])
                                time.sleep(1)
                                dt       = datetime.datetime.now()
                                out_path = save_merged(serp_bytes, land_bytes, dt)
                                append_row(ws, wb, kw, TEST_CITY,
                                           ad["heading"], ad["description"],
                                           ad["source_url"], dest_url, out_path, dt)
                                total += 1
                                print(f"  ✅ Ad {idx+1}: {ad['heading'][:50]}")
                            except Exception as e:
                                print(f"  ⚠️  Ad error: {e}")
                                if len(driver.window_handles) > 1:
                                    driver.switch_to.window(driver.window_handles[0])
                    else:
                        print(f"  ℹ️  No relevant ads on page {page_num}")

                    try:
                        nxt = driver.find_element(By.CSS_SELECTOR,
                            "a#pnnext, a[aria-label='Next page']")
                        nxt.click(); time.sleep(3)
                    except NoSuchElementException:
                        break

            except Exception as e:
                print(f"  ❌ Error: {e}")

    finally:
        if driver:
            try: driver.quit()
            except Exception: pass
        if tmp_dir and os.path.exists(tmp_dir):
            try: shutil.rmtree(tmp_dir, ignore_errors=True)
            except Exception: pass

    print(f"\n{'='*55}")
    print(f"✅ TEST DONE — {total} ads captured")
    print(f"📁 Screenshots: {SS_FOLDER}")
    print(f"📊 Excel: {OUTPUT_EXCEL}")
    print(f"{'='*55}")

if __name__ == "__main__":
    run_test()
