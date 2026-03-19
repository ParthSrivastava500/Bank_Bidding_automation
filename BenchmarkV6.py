"""
Google Sponsored Ad Automation
- Searches all Google pages per city for sponsored ads matching keyword
- Captures SERP + landing page screenshots merged side by side
- Date/time watermark on bottom-right of every JPG
- Saves to Desktop/<keyword>/ folder
- Writes to Desktop/ad_results.xlsx
- Rotates through all 15 Indian cities (collects ALL ads from ALL pages per city first)
- Pauses for manual CAPTCHA solving if detected
"""
import os, time, datetime, glob, tempfile, shutil
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from PIL import Image, ImageDraw, ImageFont
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
import pyautogui

# ── CONFIG ───────────────────────────────────────────────────
CHROME_EXE        = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
CHROMEDRIVER_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "chromedriver.exe")
AWR_EXT_PATH      = r"C:\Users\sriva\AppData\Local\Google\Chrome\User Data\Profile 11\Extensions\kjffabibegojdmknibcophlonmadkpin\1.2_0"
AWR_EXTENSION_ID  = "kjffabibegojdmknibcophlonmadkpin"

INDIAN_CITIES = [
    "Mumbai","Delhi","Bangalore","Chennai","Hyderabad",
    "Kolkata","Pune","Ahmedabad","Jaipur","Surat",
    "Lucknow","Nagpur","Indore","Bhopal","Visakhapatnam",
]
DESKTOP_PATH = Path(os.path.expanduser("~/Desktop"))
OUTPUT_EXCEL = DESKTOP_PATH / "ad_results.xlsx"
MAX_PAGES    = 10

# ── CAPTCHA HANDLER ──────────────────────────────────────────
def check_captcha(driver):
    try:
        is_captcha = (
            "sorry" in driver.current_url or
            len(driver.find_elements(By.CSS_SELECTOR,
                "form#captcha-form, div.g-recaptcha, iframe[src*='recaptcha']")) > 0
        )
        if is_captcha:
            print("\n    ⚠️  CAPTCHA detected! Please solve it in the browser window.")
            input("    ✋ Press Enter here once you have solved the CAPTCHA...")
            time.sleep(2)
            return True
    except Exception:
        pass
    return False

# ── EXCEL ────────────────────────────────────────────────────
def init_excel():
    if OUTPUT_EXCEL.exists():
        wb = openpyxl.load_workbook(OUTPUT_EXCEL)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ad Results"
        headers = ["Date","Time","keyword_term","location","ad_heading",
                   "ad_description","source_url","destination_url","screenshot_path"]
        hfill  = PatternFill("solid", start_color="1F4E79")
        hfont  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin   = Side(style="thin", color="FFFFFF")
        bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill, c.font, c.alignment, c.border = hfill, hfont, halign, bdr
        for i, w in enumerate([12,10,30,15,45,55,60,60,50], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"
        wb.save(OUTPUT_EXCEL)
    return wb, ws

def append_row(ws, wb, keyword, city, heading, desc, src, dst, path, dt):
    row  = ws.max_row + 1
    fill = PatternFill("solid", start_color="DCE6F1" if row%2==0 else "FFFFFF")
    font = Font(name="Arial", size=10)
    aln  = Alignment(vertical="top", wrap_text=True)
    for col, val in enumerate([
        dt.strftime("%d/%m/%Y"),
        dt.strftime("%H:%M:%S"),
        keyword, city, heading, desc, src, dst, str(path)
    ], 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill, c.font, c.alignment = fill, font, aln
    wb.save(OUTPUT_EXCEL)
    print(f"    ✅ Excel saved — '{keyword}' / {city} / {heading[:40]}")

# ── SCREENSHOT + WATERMARK ───────────────────────────────────
SS_FOLDER = DESKTOP_PATH / "All_Screenshots"

def keyword_folder(keyword):
    SS_FOLDER.mkdir(parents=True, exist_ok=True)
    return SS_FOLDER

def add_watermark(img, dt):
    draw = ImageDraw.Draw(img)
    text = dt.strftime("%d/%m/%Y  %H:%M:%S")
    try:
        font = ImageFont.truetype("arial.ttf", 28)
    except Exception:
        try:
            font = ImageFont.truetype("C:/Windows/Fonts/arial.ttf", 28)
        except Exception:
            font = ImageFont.load_default()
    bbox = draw.textbbox((0, 0), text, font=font)
    tw = bbox[2] - bbox[0]
    th = bbox[3] - bbox[1]
    margin = 15
    x = img.width - tw - margin
    y = img.height - th - margin
    draw.text((x+2, y+2), text, font=font, fill=(0, 0, 0, 180))
    draw.text((x, y), text, font=font, fill=(255, 255, 0))
    return img

TAB_BAR_HEIGHT = 0  # No cropping needed — driver.get_screenshot_as_png() never captures tab bar

def crop_tab_bar(img):
    """Crop the Chrome tab bar from the top of the screenshot."""
    return img.crop((0, TAB_BAR_HEIGHT, img.width, img.height))

def merge_side_by_side(b1, b2, out_path, dt):
    i1 = Image.open(io.BytesIO(b1)).convert("RGB")
    i2 = Image.open(io.BytesIO(b2)).convert("RGB")
    # Crop tab bar from both images
    i1 = crop_tab_bar(i1)
    i2 = crop_tab_bar(i2)
    h  = max(i1.height, i2.height)
    if i1.height != h: i1 = i1.resize((i1.width, h), Image.LANCZOS)
    if i2.height != h: i2 = i2.resize((i2.width, h), Image.LANCZOS)
    # Add watermark to BOTH screenshots individually
    i1 = add_watermark(i1, dt)
    i2 = add_watermark(i2, dt)
    merged = Image.new("RGB", (i1.width+i2.width, h), (255,255,255))
    merged.paste(i1, (0,0))
    merged.paste(i2, (i1.width,0))
    merged.save(out_path, "JPEG", quality=90)
    print(f"    🖼️  Saved → {out_path.name}")

# ── BROWSER ──────────────────────────────────────────────────
def create_driver():
    options = Options()
    options.binary_location = CHROME_EXE
    tmp_dir = tempfile.mkdtemp(prefix="chrome_auto_")
    options.add_argument(f"--user-data-dir={tmp_dir}")
    if os.path.exists(AWR_EXT_PATH):
        options.add_argument(f"--load-extension={AWR_EXT_PATH}")
        print(f"  🔌 AWR extension loaded")
    else:
        print(f"  ⚠️  AWR not found")
    options.add_argument("--start-maximized")
    options.add_argument("--no-first-run")
    options.add_argument("--no-default-browser-check")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    service = Service(CHROMEDRIVER_PATH)
    driver  = webdriver.Chrome(service=service, options=options)
    driver.set_page_load_timeout(30)
    driver._tmp_dir = tmp_dir
    driver.maximize_window()  # Start maximized — user can minimize manually
    return driver

# ── KEYWORD RELEVANCE FILTER ─────────────────────────────────
STOPWORDS = {"apply","online","card","get","now","for","a","the","and","in",
             "of","to","how","best","top","free","new","with","my","is","on"}

def is_relevant(ad, keyword):
    brand_words = [w for w in keyword.lower().split()
                   if w not in STOPWORDS and len(w) > 2]
    combined = (ad["heading"] + " " + ad["source_url"]).lower()
    return any(w in combined for w in brand_words)

# ── AD DESCRIPTION ───────────────────────────────────────────
def get_description(container, heading):
    for sel in [".yDYNvb", "div.yDYNvb", "span.yDYNvb", ".MUxGbd",
                "[data-sncf='1']", ".lEBKkf", ".Va3FIb"]:
        try:
            el = container.find_element(By.CSS_SELECTOR, sel)
            t = el.text.strip()
            if t and len(t) > 20 and t != heading and "http" not in t and "www." not in t:
                return t
        except Exception:
            pass
    try:
        for el in container.find_elements(By.XPATH, ".//div | .//span"):
            try:
                t = el.text.strip()
                if (len(t) > 25 and t != heading
                        and "http" not in t and "www." not in t):
                    children = el.find_elements(By.XPATH, "./div")
                    if not children:
                        return t[:300]
            except Exception:
                pass
    except Exception:
        pass
    return ""

# ── AD DETECTION ─────────────────────────────────────────────
def find_ads(driver, keyword):
    ads = []
    sponsored_labels = driver.find_elements(
        By.XPATH, "//span[text()='Sponsored'] | //div[text()='Sponsored']")

    for lbl in sponsored_labels:
        try:
            container = None
            for xpath in ["ancestor::div[@data-hveid][1]",
                          "ancestor::div[.//h3][1]",
                          "ancestor::li[1]"]:
                try:
                    container = lbl.find_element(By.XPATH, xpath)
                    if container:
                        break
                except Exception:
                    pass
            if not container:
                continue
            h3_list = container.find_elements(By.CSS_SELECTOR, "h3")
            if not h3_list:
                h3_list = container.find_elements(By.CSS_SELECTOR, "[role='heading']")
            for h3 in h3_list:
                try:
                    heading = h3.text.strip()
                    if not heading or heading == "Sponsored":
                        continue
                    try:
                        link_el = h3.find_element(By.XPATH, "ancestor::a[1]")
                    except Exception:
                        link_el = container.find_element(By.CSS_SELECTOR, "a[href]")
                    href = link_el.get_attribute("href") or ""
                    desc = get_description(container, heading)
                    ad = {"heading": heading, "description": desc,
                          "source_url": href, "link_el": link_el}
                    if (is_relevant(ad, keyword)
                            and not any(a["heading"] == heading for a in ads)):
                        ads.append(ad)
                except Exception:
                    pass
        except Exception:
            pass

    if not ads:
        for c in driver.find_elements(By.CSS_SELECTOR, "div[data-text-ad]"):
            try:
                heading = ""
                for s in ["h3", "[role='heading']"]:
                    try:
                        heading = c.find_element(By.CSS_SELECTOR, s).text.strip()
                        if heading:
                            break
                    except Exception:
                        pass
                desc    = get_description(c, heading)
                link_el = c.find_element(By.CSS_SELECTOR, "a[href]")
                href    = link_el.get_attribute("href") or ""
                ad = {"heading": heading, "description": desc,
                      "source_url": href, "link_el": link_el}
                if (is_relevant(ad, keyword)
                        and not any(a["heading"] == heading for a in ads)):
                    ads.append(ad)
            except Exception:
                pass
    return ads

# ── SEARCH ALL PAGES FOR ONE CITY ────────────────────────────
def search_all_pages_in_city(driver, keyword, city, wb, ws):
    folder      = keyword_folder(keyword)
    total_found = 0

    try:
        driver.get("https://www.google.co.in/")
        time.sleep(3)
        check_captcha(driver)
        try:
            driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
            time.sleep(1)
        except Exception:
            pass
        for sel in ["button#L2AGLb", "button[aria-label*='Accept']"]:
            try:
                driver.find_element(By.CSS_SELECTOR, sel).click()
                time.sleep(1)
                break
            except Exception:
                pass
        sb = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.NAME, "q")))
        sb.clear()
        sb.send_keys(keyword)
        sb.send_keys(Keys.RETURN)
        time.sleep(3)
        check_captcha(driver)
    except Exception as e:
        print(f"    ❌ Google failed: {e}")
        return 0

    for page_num in range(1, MAX_PAGES+1):
        print(f"    📄 Page {page_num} — scanning...")
        ads = find_ads(driver, keyword)

        if ads:
            print(f"    🎯 {len(ads)} relevant ad(s) on page {page_num}")
            timestamp = datetime.datetime.now()
            ts_str    = timestamp.strftime("%Y%m%d_%H%M%S")
            for idx, ad in enumerate(ads):
                try:
                    print(f"    🔗 Ad {idx+1}: {ad['heading'][:60]}")
                    driver.execute_script(
                        "arguments[0].scrollIntoView({block: 'center'});", ad["link_el"])
                    time.sleep(1)
                    # Capture full window then crop ONLY the tab bar (top 35px)
                    _raw = pyautogui.screenshot()
                    _buf = io.BytesIO(); _raw.save(_buf, 'PNG')
                    _img = Image.open(io.BytesIO(_buf.getvalue())).convert("RGB")
                    _img = _img.crop((0, 35, _img.width, _img.height))
                    _out = io.BytesIO(); _img.save(_out, 'PNG')
                    serp_bytes = _out.getvalue()
                    driver.execute_script(
                        "arguments[0].setAttribute('target','_blank');", ad["link_el"])
                    ad["link_el"].click()
                    import random as _r2
                    time.sleep(_r2.uniform(2, 4))
                    if len(driver.window_handles) > 1:
                        driver.switch_to.window(driver.window_handles[-1])
                        time.sleep(3)
                    dest_url   = driver.current_url
                    # Capture full window then crop ONLY the tab bar (top 35px)
                    _raw2 = pyautogui.screenshot()
                    _buf2 = io.BytesIO(); _raw2.save(_buf2, 'PNG')
                    _img2 = Image.open(io.BytesIO(_buf2.getvalue())).convert("RGB")
                    _img2 = _img2.crop((0, 35, _img2.width, _img2.height))
                    _out2 = io.BytesIO(); _img2.save(_out2, 'PNG')
                    land_bytes = _out2.getvalue()
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                    time.sleep(1)
                    safe_h   = "".join(c if c.isalnum() else "_" for c in ad["heading"][:40])
                    out_path = folder / f"{ts_str}_p{page_num}_ad{idx+1}_{safe_h}.jpg"
                    merge_side_by_side(serp_bytes, land_bytes, out_path, timestamp)
                    append_row(ws, wb, keyword, city,
                               ad["heading"], ad["description"],
                               ad["source_url"], dest_url, out_path, timestamp)
                    total_found += 1
                except Exception as e:
                    print(f"    ⚠️  Ad {idx+1} error: {e}")
                    if len(driver.window_handles) > 1:
                        driver.switch_to.window(driver.window_handles[0])
        else:
            print(f"    ℹ️  No relevant ads on page {page_num}")

        try:
            nxt = driver.find_element(By.CSS_SELECTOR,
                "a#pnnext, a[aria-label='Next page']")
            nxt.click()
            import random as _r
            page_delay = _r.uniform(2, 4)
            time.sleep(page_delay)
            check_captcha(driver)
        except NoSuchElementException:
            print(f"    ℹ️  No more pages after page {page_num}.")
            break

    return total_found

# ── PER-KEYWORD RUNNER ───────────────────────────────────────
def process_keyword(keyword):
    print(f"\n{'='*60}\n🔍 Keyword: '{keyword}'\n{'='*60}")
    wb, ws  = init_excel()
    driver  = None
    tmp_dir = None
    grand_total = 0
    try:
        driver  = create_driver()
        tmp_dir = getattr(driver, '_tmp_dir', None)
        time.sleep(2)
        print("  ✅ Browser ready")
        for city in INDIAN_CITIES:
            print(f"\n  📍 City: {city} — scanning all pages...")
            # Clear cookies between cities to reduce CAPTCHA risk
            try:
                driver.delete_all_cookies()
                driver.execute_script("window.localStorage.clear(); window.sessionStorage.clear();")
                time.sleep(1)
            except Exception:
                pass
            count = search_all_pages_in_city(driver, keyword, city, wb, ws)
            grand_total += count
            # Random delay between cities to look human
            import random
            city_delay = random.randint(3, 5)
            print(f"  ⏳ Waiting {city_delay}s before next city...")
            time.sleep(city_delay)
            print(f"\n  📊 {count} ads captured in {city}. Total so far: {grand_total}")
        print(f"\n  🏁 Finished all cities. Total ads captured: {grand_total}")
    except Exception as e:
        print(f"\n  ❌ Fatal: {e}")
    finally:
        if driver:
            try: driver.quit()
            except Exception: pass
        if tmp_dir and os.path.exists(tmp_dir):
            try: shutil.rmtree(tmp_dir, ignore_errors=True)
            except Exception: pass

# ── MAIN ─────────────────────────────────────────────────────
def main():
    print("="*60)
    print("  Google Sponsored Ad Automation")
    print(f"  Screenshots → Desktop/<keyword>/")
    print(f"  Excel       → {OUTPUT_EXCEL}")
    print("="*60)
    print("\nEnter keywords (comma separated) or 'quit' to exit.\n")
    init_excel()
    while True:
        try:
            raw = input("🔑 Enter keyword(s): ").strip()
        except (KeyboardInterrupt, EOFError):
            print("\nGoodbye!")
            break
        if not raw: continue
        if raw.lower() in ("quit","exit","q"):
            print("Goodbye!")
            break
        keywords = [k.strip() for k in raw.split(",") if k.strip()]
        if not keywords: continue
        print(f"\n▶ Running {len(keywords)} keyword(s) in parallel...\n")
        if len(keywords) == 1:
            process_keyword(keywords[0])
        else:
            with ThreadPoolExecutor(max_workers=min(len(keywords),4)) as ex:
                futures = {ex.submit(process_keyword,kw):kw for kw in keywords}
                for f in as_completed(futures):
                    try: f.result()
                    except Exception as e: print(f"❌ {e}")
        print(f"\n✅ Done!\n{'-'*60}")

if __name__ == "__main__":
    main()